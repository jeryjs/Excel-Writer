from fnmatch import translate
import json
import os
import sys
import openpyxl
from openpyxl.utils import column_index_from_string

def update_cells(ws, mappings):
    for cell, value in mappings.items():
        try:
            ws[cell] = float(value)
        except ValueError:
            ws[cell] = value

def remove_column(ws, col, count):
    col_idx = column_index_from_string(col)
    ws.delete_cols(col_idx, count)
    ws.move_range(f"{col}1:{col}1000", rows=-count, cols=0, translate=True)

def insert_column(ws, col, count):
    col_idx = column_index_from_string(col)
    ws.insert_cols(col_idx, count)
    ws.move_range(f"{col}1:{col}1000", rows=0, cols=count, translate=True)

def main(args):
    # If operations was not provided, use some default values
    if len(args) < 2 or args[1] == "":
        # default operations...
        operations = [
            {"type": "updateCells", "sheet": "START", "mappings": {"C06": "Test", "C07": "Test Position", "C08": "GoLang", "C09": "22GO01", "C10": "CSE/AI", "C11": "4", "C12": "2024"}},
            {"type": "insertColumn", "sheet": "IA", "column": "N", "count": 3},
            # {"type": "removeColumn", "sheet": "IA", "column": "R"},
			{"type": "updateCells", "sheet": "IA", "mappings": {"E08": "CO1", "F08": "CO2", "G08": "CO3", "H08": "CO4", "I08": "CO5", "J08": "CO6", "K08": "CO1", "L08": "CO2", "M08": "CO3", "N08": "CO4", "O08": "CO5"}},
			{"type": "updateCells", "sheet": "IA", "mappings": {"E09": "5", "F09": "5", "G09": "5", "H09": "5", "I09": "5", "J09": "5", "K09": "5", "L09": "5", "M09": "5", "N09": "5", "O09": "5"}},
			{"type": "updateCells", "sheet": "IA", "mappings": {"E10": "3", "F10": "3", "G10": "3", "H10": "3", "I10": "3", "J10": "3", "K10": "3", "L10": "3", "M10": "3", "N10": "3", "O10": "3"}},
			{"type": "updateCells", "sheet": "IA", "mappings": {"E11": "4", "F11": "4", "G11": "4", "H11": "4", "I11": "4", "J11": "4", "K11": "4", "L11": "4", "M11": "4", "N11": "4", "O11": "4"}},
			{"type": "updateCells", "sheet": "IA", "mappings": {"E12": "5", "F12": "5", "G12": "5", "H12": "5", "I12": "5", "J12": "5", "K12": "5", "L12": "5", "M12": "5", "N12": "5", "O12": "5"}},
        ]
        args.append(json.dumps(operations))
        if len(args) < 3:
            args.append("input2.xlsx")

    # Unmarshal the JSON input
    operations = json.loads(args[1])

    # If input file path was provided use it, else read from standard input
    wb = openpyxl.load_workbook(args[2])

    # Perform the operations on the Excel file
    for op in operations:
        ws = wb[op['sheet']]
        if op['type'] == 'updateCells':
            update_cells(ws, op['mappings'])
        elif op['type'] == 'removeColumn':
            remove_column(ws, op['column'], op['count'])
            col_idx = column_index_from_string(op['column'])
            ws.delete_cols(col_idx, op['count'])

        elif op['type'] == 'insertColumn':
            insert_column(ws, op['column'], op['count'])
        else:
            print(f"Unknown operation type: {op['type']}", file=sys.stderr)

    # Save the workbook
    wb.save('output.xlsx')

if __name__ == "__main__":
    main(sys.argv)
    os.system('output.xlsx')